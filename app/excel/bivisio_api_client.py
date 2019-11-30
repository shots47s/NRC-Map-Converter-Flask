
import requests
import os
import json
import pyexcel as p
from collections import OrderedDict
import pprint
import datetime
import openpyxl
from io import BytesIO

nrc2bivisio_main_map_dict = OrderedDict([
    ('id', 'Id company'),
    ('name', 'company.name'),

])

nrc2bivisio_branch_map_dict = OrderedDict([
    ('address', 'address_line_1'),
    ('city', 'city'),
    ('state', 'state'),
    ('country', 'country'),
    ('postal_code', 'postal_code'),
    ('website', 'website')
])

nrc2bivisio_map_dict = OrderedDict([
    #('id','Id company'),
    # ('name','company.name'),
    # ()



    ('Main Site', 'Id Branch'),
    ('Parent', 'Main Site'),
    ('company.name', 'company.name'),
    ('Website', 'Website'),
    ('NAICS', 'Categories'),
    ('address_line_1', 'address_line_1'),
    ('city', 'city'),
    ('state', 'state'),
    ('Prov_Original', 'Provinces'),
    ('country', 'country'),
    ('postal_code', 'postal_code'),
    ('Influence', 'Influences'),
    ('Type', 'Types'),
    ('Sector', 'Sectors'),
    ('Enabler/Processing', 'Categories'),
    ('Theme', 'Themes'),
    ('Description', 'Description_EN-1'),
    ('Email', 'Email')
])

canonical_nrc_headers = [
    'Field',
    'Id company',
    'company.name',
    'alternative search terms ',
    'Invitation.Contact',
    'Date constitution',
    'Sales_Revenue',
    'Description_Fr',
    'Description_EN',
    'Nb employés – Admin',
    'Nb employés – Prod',
    'Nb employés – Recherche',
    'Id Branch',
    'Main Site',
    'Parent',
    'branch.name',
    'Website',
    'address_line_1',
    'city',
    'state',
    'country',
    'postal_code',
    'Telephone',
    'Description_fr',
    'Description_EN-1',
    'Email',
    'Sectors',
    'Themes',
    'Influences',
    'Provinces',
    'Types',
    'Categories',
    'NAICS2017'
]

def get_safe_ws_value(ws,headers,row_num,column_name):
        if column_name not in headers:
            raise RuntimeError("Column Name: {} not in the NRC spreadsheet".format(column_name))

        # Kludge, their column names are not unique
        column_num = headers.index(column_name)
        if column_name == "Description_EN":
            column_num = 25

        cell = ws.cell(row=row_num,column=column_num)

        if cell.value is None:
            return ""

        if cell.hyperlink is not None:
            if cell.hyperlink.display is not None:
                return cell.hyperlink.display

        return cell.value

class Bivzio_API_Client:
    def __init__(self, _token=None, _url=None):
        self.token = _token or os.environ.get('BIVIZIO_API_TOKEN')
        self.url = _url or os.environ.get('BIVIZIO_API_URL')
        if self.token is None:
            raise RuntimeError("There is no API Token for the BIVIZIO API,"
                               "please set the environment variable BIVIZIO_API_TOKEN")

    def get_from_id(self, _id=-1, use_biv_id=True):
        get_url = self.url + "/company/get/{}/".format(_id)

        if use_biv_id:
            params = {'token': self.token}
        else:
            params = {'token': self.token,
                      'use_external_id': 'true'}

        try:
            r = requests.get(get_url,
                             params=params,
                             verify=False)

            if r.json() == {'message': 'No company matching ID'}:
                return None

            return r.json()

        except requests.exceptions.RequestException as e:
            raise RuntimeError("Bivizio API Error: {}".format(e))

    def create_new_entry(self, biv_json_):
        ### Ideally, we would have a schema to validate against, but for now will wrap in try

        create_url = "{}/company/create/?token={}".format(self.url,self.token)

        json_string = json.dumps(biv_json_)

        print(json_string)
        try:
            r = requests.post(create_url,
                              headers = {'Content-type':'application/json'},
                              data=json_string,
                              verify=False)
            r.raise_for_status()

        except requests.exceptions.RequestException as e:
            raise RuntimeError("Bivizio API Error: {}".format(e))

    def edit_exist_entry(self, external_id, biv_json_update):
        ### This will be an edit function that will only update

        #get existing json
        exist_json = self.get_from_id(external_id,use_biv_id=False)
        #pp.pprint(exist_json)
        #print(json.dumps(exist_json))
        if exist_json is None:
            raise RuntimeError("Trying to Edit a non existing entry {}".format(external_id))

        edit_url = "{}/company/edit/{}/?token={}&use_external_id=true".format(self.url, external_id,self.token)
        json_string = json.dumps(biv_json_update)
        print(edit_url)
        try:
            r = requests.post(edit_url,
                              headers = {'Content-type':'application/json'},
                              data=json_string,
                              verify=False)
            r.raise_for_status()
        except requests.exceptions.RequestException as e:
            raise RuntimeError("Bivizio API Error: {}".format(e))

    @staticmethod
    def convert_nrc_workbook_to_biv_json_list(_nrc_file):
        # First create a map of all of the rows that need to be branchs
        row_branch_map = {}

        # Get Headers for index lookup
        wb = openpyxl.load_workbook(filename=BytesIO(_nrc_file))
        ws = wb.active

        # get headers for indexing
        headers = [x.value for x in ws['A1':'AH1'][0]]
        ## Add a blank space so I don't have to add 1 to every index call
        ## excel starts indexing at 1
        headers.insert(0,"")
        print(headers)
        print(headers.index('Id company'))

        # first populate the main sites and then run through again to get branches
        for ir in range(3,ws.max_row):
            if get_safe_ws_value(ws,headers,ir,'Main Site'):
                id_of_company = get_safe_ws_value(ws,headers,ir,'Id company')
                #print("{}".format(id_of_company))
                row_branch_map[id_of_company] = {'row_num': ir, 'branches':[] }


        for ir in range(3,ws.max_row):
            if get_safe_ws_value(ws,headers,ir,'Main Site') != 1:
                parent_id = get_safe_ws_value(ws,headers,ir,'Parent')
                id_of_company = get_safe_ws_value(ws,headers,ir,'Id company')
                ws.cell(row=ir, column=headers.index('Id company')).value
                #print("{} {}".format(parent_id,id_of_company))
                row_branch_map[parent_id]['branches'].append((ir,id_of_company))



        # nrc_rows = p.get_records(file_name=_nrc_file)

        # # Correct Description_EN: KLUDGE
        # for row in nrc_rows:
        #     row['Description_EN'] = row['Description_EN-1']

        # # first populate the main sites and then run through again to get branches
        # for row in nrc_rows[1:]:
        #     if row['Main Site']:
        #         id_of_company = row['Id company']
        #         row_branch_map[id_of_company] = {
        #             'row_num': nrc_rows.index(row), 'branches': []}

        # for row in nrc_rows[1:]:
        #     if row['Main Site'] != 1:
        #         parent_id = row['Parent']
        #         id_of_company = row['Id company']
        #         row_branch_map[parent_id]['branches'].append(
        #             (nrc_rows.index(row), id_of_company))

        # now that we have the map, lets make the list of jsons
        todays_date_string = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        #biv_id = 1

        biv_json_array = []

        for id, info in row_branch_map.items():
            main_company_row_num = info['row_num']
            # place main information
            biv_json = {#'id': "{}".format(main_company_row_num),
                        'status': "1",
                        'date_created': todays_date_string,  # make their format
                        'date_formed': '0000-00-00',
                        'name': get_safe_ws_value(ws,headers,main_company_row_num,'company.name'),
                        'sales': get_safe_ws_value(ws,headers,main_company_row_num,'Sales_Revenue'),
                        'description_fr': "",
                        'description_en': "",
                        'employees_admin': get_safe_ws_value(ws,headers,main_company_row_num,'Nb employés – Admin'),
                        "employees_prod": get_safe_ws_value(ws,headers,main_company_row_num,'Nb employés – Prod'),
                        "employees_research": get_safe_ws_value(ws,headers,main_company_row_num,'Nb employés – Recherche'),
                        "group_id": "2",
                        "external_id": "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'Id company')),
                        "external_token": "fake_entry",
                        "internal_token": "",
                        "admin_user_id": "0",
                        "branches": [{'id': "{}".format(main_company_row_num),
                                       'status': "1",
                                       'is_main': "1",
                                       'date_created': todays_date_string,
                                       'company_id': "{}".format(main_company_row_num),
                                       'name': get_safe_ws_value(ws,headers,main_company_row_num,'company.name'),
                                       'address': get_safe_ws_value(ws,headers,main_company_row_num,'address_line_1'),
                                       'city': get_safe_ws_value(ws,headers,main_company_row_num,'city'),
                                       'state': get_safe_ws_value(ws,headers,main_company_row_num,'state').upper(),
                                       'country': get_safe_ws_value(ws,headers,main_company_row_num,'country'),
                                       'postal_code': "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'postal_code')),
                                       'description_fr': "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'Description_fr')),
                                       'description_en': "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'Description_EN')),
                                       'phone': "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'Telephone')),
                                       'email': "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'Email')),
                                       'website': "{}".format(get_safe_ws_value(ws,headers,main_company_row_num,'Website')),
                                       'lat': "",
                                       'lon': "",
                                       'mcgill': [{
                                           'naics': get_safe_ws_value(ws,headers,main_company_row_num,'NAICS2017'),
                                           'province': get_safe_ws_value(ws,headers,main_company_row_num,'Provinces').upper(),
                                           'influence': get_safe_ws_value(ws,headers,main_company_row_num,'Influences'),
                                           'type': get_safe_ws_value(ws,headers,main_company_row_num,'Types'),
                                           'sector': get_safe_ws_value(ws,headers,main_company_row_num,'Sectors'),
                                           'enabler_processing': get_safe_ws_value(ws,headers,main_company_row_num,'Categories'),
                                           'theme': get_safe_ws_value(ws,headers,main_company_row_num,'Themes')
                                       }]
                                      }]
                        }
            # Now loop through and create the branch information
            for b_row_num, b_id in info['branches']:
                br_json = {'id': b_row_num,
                           'status': 1,
                           'is_main': 0,
                           'date_created': todays_date_string,
                           'company_id': main_company_row_num,
                           'name': get_safe_ws_value(ws,headers,b_row_num,'company.name'),
                           'address': get_safe_ws_value(ws,headers,b_row_num,'address_line_1'),
                           'city': get_safe_ws_value(ws,headers,b_row_num,'city'),
                           'state': get_safe_ws_value(ws,headers,b_row_num,'state').upper(),
                           'country': get_safe_ws_value(ws,headers,b_row_num,'country'),
                           'postal_code': get_safe_ws_value(ws,headers,b_row_num,'postal_code'),
                           'description_fr': "{}".format(get_safe_ws_value(ws,headers,b_row_num,'Description_fr')),
                           'description_en': "{}".format(get_safe_ws_value(ws,headers,b_row_num,'Description_EN')),
                           'phone': get_safe_ws_value(ws,headers,b_row_num,'Telephone'),
                           'email': get_safe_ws_value(ws,headers,b_row_num,'Email'),
                           'website': get_safe_ws_value(ws,headers,b_row_num,'Website'),
                           'lat': '',
                           'lon': '',
                           'mcgill': [{
                               'naics': get_safe_ws_value(ws,headers,b_row_num,'NAICS2017'),
                               'province': get_safe_ws_value(ws,headers,b_row_num,'Provinces').upper(),
                               'influence': get_safe_ws_value(ws,headers,b_row_num,'Influences'),
                               'type': get_safe_ws_value(ws,headers,b_row_num,'Types'),
                               'sector': get_safe_ws_value(ws,headers,b_row_num,'Sectors'),
                               'enabler_processing': get_safe_ws_value(ws,headers,b_row_num,'Categories'),
                               'theme': get_safe_ws_value(ws,headers,b_row_num,'Themes')
                           }]
                           }
                biv_json['branches'].append(br_json)

            biv_json_array.append(biv_json)
        return biv_json_array

    def update_bivisio_entry(self, biv_json_):
        # Determine if this entry is in the database
        print("looking_for_biv_id: {}".format(biv_json_['external_id']))
        db_entry = self.get_from_id(biv_json_['external_id'],use_biv_id=False)
        if db_entry:
            self.edit_exist_entry(biv_json_['external_id'],biv_json_)
        else:
            self.create_new_entry(biv_json_)

    def update_bivisio_database_from_nrc_spreadsheet(self, nrc_file_name):

        json_array = self.convert_nrc_workbook_to_biv_json_list(nrc_file_name)
        print(json_array[0])
        #for biv_json in json_array:
        #    self.update_bivisio_entry(biv_json)

